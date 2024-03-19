import tkinter as tk
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import ttk
from tkinter import *

root = tk.Tk()
root.geometry("500x600")
root.title('LogIN')

excel_con = Workbook()
excel_con = load_workbook('accounts.xlsx')
excel_activate = excel_con.active

def loginfunction():
    user = username.get()
    passw = password.get()
    found = False

    if user == "" or passw == "":
        messagebox.showinfo("Notification", "Fill All Required")
    else:
        for each_cell in range(2, excel_activate.max_row + 1):
            if (username.get() == excel_activate['A' + str(each_cell)].value and password.get() == excel_activate['B' + str(each_cell)].value):
                found = True
                break
        if found:
            bank = Toplevel()
            bank.geometry("1200x700")
            bank.title("Kennedy's Online Bank")
            root.withdraw()

            bankImg = PhotoImage(file="bank.png")
            bankImg = bankImg.subsample(10, 10)

            profImg = PhotoImage(file="profile.png")
            profImg = profImg.subsample(1, 1)

            logoutImg = PhotoImage(file="logout.png")
            logoutImg = logoutImg.subsample(8, 8)

            login_frame = Frame(bank, width=1200, height=700, borderwidth=28, relief='groove', bg='lightblue')
            login_frame.place(x=0, y=0)

            bankbg_label = Label(login_frame, bg='lightblue',image=bankImg)
            bankbg_label.place(x=10, y=10)

            bankbg_label = Label(login_frame, text="Kennedy's Online Bank", font=('courier', 25, 'underline'), bg='lightblue')
            bankbg_label.place(x=140, y=50)

            prof_button = Button(login_frame, image=profImg,bg='lightblue',command=lambda:profile_function())
            prof_button.place(x=950, y=20)

            log_button = Button(login_frame, image=logoutImg, bg='lightblue',command=lambda:logout_function())
            log_button.place(x=1050, y=20)

            def back_to_login():
                bank.destroy()
                root.deiconify()
                username.delete(0,END)
                password.delete(0,END)

            switch = Button(login_frame,text='Switch Account', bg='lightblue', font=('courier', 20),command=lambda:back_to_login())
            switch.place(x=20, y=150)
            
            info_frame = Frame(login_frame,width=300,height=400, borderwidth=10, relief='groove',)
            info_frame.place(x=20,y=220)

            tree = ttk.Treeview(info_frame,height=17)
            treescrolly = Scrollbar(info_frame, orient="vertical", command= tree.yview)
            tree.configure( yscrollcommand=treescrolly.set) 
            treescrolly.place(x=260,y=0)
            tree['columns'] = ('User Id','Amount','Type')

            tree.column('#0',width=0,stretch=0)
            tree.column('User Id',anchor='center',width=70)
            tree.column('Amount',anchor='center',width=90)
            tree.column('Type',anchor='center',width=95)

            tree.heading('User Id',text='User Id')
            tree.heading('Amount',text='Amount')
            tree.heading('Type',text='Type')
            for each_cell in range(2, (excel_activate.max_row)+1):
                tree.insert(parent='', index="end", values=(excel_activate['A'+str(each_cell)].value,excel_activate['C'+str(each_cell)].value, excel_activate['D'+str(each_cell)].value))
            tree.place(x=0,y=0)

            all_frame = Frame(login_frame,width=790,height=510, borderwidth=10, relief='groove', bg='lightblue')
            all_frame.place(x=340,y=120)

            def withdraw_function():
                withdraw1 = Toplevel()
                withdraw1.geometry("300x200")
                withdraw1.title('WITHDRAW AMOUNT')
                withdraw1.configure(bg='lightblue')
                bank.withdraw()

                excel_con = load_workbook('accounts.xlsx')
                excel_activate = excel_con.active

                def refresh_data(tree):
                    tree.delete(*tree.get_children())
                    data = get_updated_data()
                    for item in data:
                        tree.insert('', 'end', values=item)

                def get_updated_data():
                    updated_value = list()
                    for each_cell in range(2, (excel_activate.max_row) + 1):
                        updated_value.append([excel_activate['A' + str(each_cell)].value,
                                            excel_activate['C' + str(each_cell)].value,
                                            excel_activate['D' + str(each_cell)].value])
                    return updated_value
                            

                def on_entry_click1(event):
                    if withdraw_ent.get() == 'Enter UserID':
                        withdraw_ent.delete(0, tk.END)  # Remove the default text on entry click
                        withdraw_ent.config(fg='black') # Change the text color to black

                def withdraw_function():
                    user_id = withdraw_ent.get()  # Get the user ID from the Entry widget
                    if user_id == "":
                        messagebox.showinfo("Notif", "User ID")
                    else:
                        for each_cell in range(2, (excel_activate.max_row) + 1):
                            if user_id == excel_activate['A' + str(each_cell)].value:
                                found = True
                                break
                            else:
                                found = False
                        if found :
                            withdraw = tk.Toplevel()
                            withdraw.geometry("300x300")
                            withdraw.title("Withdraw Window")
                            withdraw1.destroy()

                            def perform_withdraw():
                                withdraw_value = withdraw_entry.get()
                                if withdraw_value and withdraw_value.isdigit():
                                    current_value = excel_activate['C' + str(each_cell)].value
                                    if current_value is not None:
                                        current_value = int(current_value)
                                    else:
                                        current_value = 0

                                    new_value = current_value - int(withdraw_value)
                                    excel_activate['C' + str(each_cell)].value = new_value
                                    excel_con.save('accounts.xlsx')
                                    messagebox.showinfo("Notification", "Depo Completed")
                                    refresh_data(tree)
                                    withdraw.destroy()
                                    bank.deiconify()
                                else:
                                    messagebox.showerror("Error", "Invalid deposit amount. Please enter a valid number.")

                            withdraw_amountlbl = tk.Label(withdraw, text='Current Amount', font=('arial', 15, 'underline'))
                            withdraw_amountlbl.pack()

                            withdraw_amount = tk.Label(withdraw, text=excel_activate['C' + str(each_cell)].value, font=('arial', 15))
                            withdraw_amount.pack()

                            def on_entry_click(event):
                                if withdraw_entry.get() == 'Enter Amount':
                                    withdraw_entry.delete(0, tk.END)  # Remove the default text on entry click
                                    withdraw_entry.config(fg='black')  # Change the text color to black

                            withdraw_entry = tk.Entry(withdraw, width=13, font=("arial", 20))
                            withdraw_entry.insert(0, 'Enter Amount')
                            withdraw_entry.bind('<FocusIn>', on_entry_click)
                            withdraw_entry.pack()

                            withdraw_btn = tk.Button(withdraw, text="Withdraw", font=('arial', 20), command=perform_withdraw, bg='lightblue')
                            withdraw_btn.pack()
                                                
                        else:
                            messagebox.showerror('Error','Wrong User ID')

                withdraw_frame = Frame(withdraw1,width=300,height=200,borderwidth=20,relief='groove')
                withdraw_frame.place(x=0,y=0)

                withdraw_ent = Entry(withdraw_frame,width=13,font=("arial",20))
                withdraw_ent.insert(0, 'Enter UserID')
                withdraw_ent.bind('<FocusIn>', on_entry_click1)
                withdraw_ent.place(x=10,y=10)

                withdraw_btn = Button(withdraw_frame,text="Enter",font=('arial',20),command=lambda:withdraw_function(),bg='lightblue')
                withdraw_btn.place(x=10,y=50)


                withdraw1.mainloop()

            def deposit_function():
                deposit1 = Toplevel()
                deposit1.geometry("300x200")
                deposit1.title('DEPOSIT AMOUNT')
                deposit1.configure(bg='lightblue')
                bank.withdraw()

                excel_con = load_workbook('accounts.xlsx')
                excel_activate = excel_con.active

                def on_entry_click1(event):
                    if deposit_ent.get() == 'Enter UserID':
                        deposit_ent.delete(0, tk.END)  # Remove the default text on entry click
                        deposit_ent.config(fg='black') # Change the text color to black

                def deposit_function():
                    user_id = deposit_ent.get() 
                    if user_id == "":
                        messagebox.showinfo("Notif", "User ID")
                    else:
                        for each_cell in range(2, (excel_activate.max_row) + 1):
                            if user_id == excel_activate['A' + str(each_cell)].value:
                                found = True
                                break
                            else:
                                found = False
                        if found :
                            deposit = tk.Toplevel()
                            deposit.geometry("300x300")
                            deposit.title("Deposit Window")
                            deposit1.destroy()

                            def refresh_data(tree):
                                tree.delete(*tree.get_children())
                                data = get_updated_data()
                                for item in data:
                                    tree.insert('', 'end', values=item)

                            def get_updated_data():
                                updated_value = list()
                                for each_cell in range(2, (excel_activate.max_row) + 1):
                                    updated_value.append([excel_activate['A' + str(each_cell)].value,
                                                        excel_activate['C' + str(each_cell)].value,
                                                        excel_activate['D' + str(each_cell)].value])
                                return updated_value
                            
                            def perform_deposit():
                                deposit_value = deposit_entry.get()
                                if deposit_value and deposit_value.isdigit():
                                    current_value = excel_activate['C' + str(each_cell)].value
                                    if current_value is not None:
                                        current_value = int(current_value)
                                    else:
                                        current_value = 0

                                    new_value = current_value + int(deposit_value)
                                    excel_activate['C' + str(each_cell)].value = new_value
                                    excel_con.save('accounts.xlsx')
                                    messagebox.showinfo("Notification", "Depo Completed")
                                    refresh_data(tree)
                                    deposit.destroy()
                                    bank.deiconify()
                                else:
                                    messagebox.showerror("Error", "Invalid deposit amount. Please enter a valid number.")                                
                            deposit_amountlbl = tk.Label(deposit, text='Current Amount', font=('arial', 15, 'underline'))
                            deposit_amountlbl.pack()

                            deposit_amount = tk.Label(deposit, text=excel_activate['C' + str(each_cell)].value, font=('arial', 15))
                            deposit_amount.pack()

                            def on_entry_click(event):
                                if deposit_entry.get() == 'Enter Amount':
                                    deposit_entry.delete(0, tk.END)  # Remove the default text on entry click
                                    deposit_entry.config(fg='black')  # Change the text color to black

                            deposit_entry = tk.Entry(deposit, width=13, font=("arial", 20))
                            deposit_entry.insert(0, 'Enter Amount')
                            deposit_entry.bind('<FocusIn>', on_entry_click)
                            deposit_entry.pack()

                            deposit_btn = tk.Button(deposit, text="deposit", font=('arial', 20), command=perform_deposit, bg='lightblue')
                            deposit_btn.pack()

                            deposit.mainloop()
                                                
                        else:
                            messagebox.showerror('Error','Wrong User ID')

                deposit_frame = Frame(deposit1,width=300,height=200,borderwidth=20,relief='groove')
                deposit_frame.place(x=0,y=0)

                deposit_ent = Entry(deposit_frame,width=13,font=("arial",20))
                deposit_ent.insert(0, 'Enter UserID')
                deposit_ent.bind('<FocusIn>', on_entry_click1)
                deposit_ent.place(x=10,y=10)

                deposit_btn = Button(deposit_frame,text="Enter",font=('arial',20),command=lambda:deposit_function(),bg='lightblue')
                deposit_btn.place(x=10,y=50)


                deposit1.mainloop()

            depoImg = PhotoImage(file="deposit.png")
            depoImg = depoImg.subsample(1, 1)

            withImg = PhotoImage(file="withdraw.png")
            withImg = withImg.subsample(1, 1)

            myinfoImg = PhotoImage(file="myinfo.png")
            myinfoImg = myinfoImg.subsample(1, 1)

            depobutton = Button(all_frame, image=depoImg, bg='lightblue',command=lambda:deposit_function())
            depobutton.place(x=320,y=50)

            withbutton = Button(all_frame, image=withImg, bg='lightblue',command=lambda:withdraw_function())
            withbutton.place(x=570,y=50)

            myinfobutton = Button(all_frame, image=myinfoImg, bg='lightblue',command=lambda:profile_function())
            myinfobutton.place(x=50,y=50)

            def logout_function():
                logout = Toplevel()
                logout.geometry('200x200')
                logout.title('Option')
                logout.configure(bg='lightblue')
                bank.destroy()

                def back_to_login():
                    logout.destroy()
                    root.deiconify()
                    username.delete(0,END)
                    password.delete(0,END)
                def exit():
                    logout.destroy()
                back_to_login_btn = Button(logout,text="Sign Out",font=('arial',30,'underline'),command=lambda:back_to_login(),bg='lightblue')
                back_to_login_btn.pack(fill=X)

                exit_btn = Button(logout,text="Exit",font=('arial',30,'underline'),command=lambda:exit(),bg='lightblue')
                exit_btn.pack(fill=X)
            
                logout.mainloop()

            def profile_function():
                profilem = Toplevel()
                profilem.geometry('300x200')
                profilem.title('YOUR PROFILE')
                bank.withdraw()

                def refresh_data(tree):
                    tree.delete(*tree.get_children())
                    data = get_updated_data()
                    for item in data:
                        tree.insert('', 'end', values=item)

                def get_updated_data():
                    updated_value = list()
                    for each_cell in range(2, (excel_activate.max_row) + 1):
                        updated_value.append([excel_activate['A' + str(each_cell)].value,
                                            excel_activate['C' + str(each_cell)].value,
                                            excel_activate['D' + str(each_cell)].value])
                    return updated_value


                profile_frame = Frame(profilem,width=300,height=600,borderwidth=20,relief='groove')
                profile_frame.place(x=0,y=0)
                
                def on_entry_click1(event):
                    if profilem_ent.get() == 'Enter UserID':
                        profilem_ent.delete(0, tk.END)  # Remove the default text on entry click
                        profilem_ent.config(fg='black') # Change the text color to black

                def profilem_function():
                    user_id = profilem_ent.get()  # Get the user ID from the Entry widget
                    if user_id == "":
                        messagebox.showinfo("Notif", "User ID")
                    else:
                        for each_cell in range(2, (excel_activate.max_row) + 1):
                            if user_id == excel_activate['A' + str(each_cell)].value:
                                found = True
                                cell_address = each_cell
                                break
                            else:
                                found = False
                        if found :
                            profile = Toplevel()
                            profile.geometry('300x400')
                            profile.title('YOUR PROFILE')
                            profilem.destroy()

                            def update_combobox_value():
                                type_var.set(excel_activate['D' + str(each_cell)].value)

                            def go_back_to_main(bank, profile):
                                profile.destroy()
                                bank.deiconify()

                            def save():
                                excel_activate['A'+str(each_cell)].value = proftext.get()
                                excel_activate['D'+str(each_cell)].value = type_var.get()
                                excel_con.save('accounts.xlsx')
                                refresh_data(tree)
                                messagebox.showinfo("Notif","Save Successfull")

                            def delete():
                                excel_activate.delete_rows(cell_address)
                                messagebox.showinfo("Delete","Account Deleted")
                                excel_con.save('accounts.xlsx')
                                refresh_data(tree)
                                profile.destroy()
                                bank.deiconify()
                                return

                            profImg = PhotoImage(file="profile.png")
                            profImg = profImg.subsample(1, 1)

                            profile_frame = Frame(profile,width=300,height=400,borderwidth=20,relief='groove',bg='lightblue')
                            profile_frame.place(x=0,y=0)

                            profimg = Label(profile_frame,text="",image=profImg,bg='lightblue')
                            profimg.place(x=15,y=15)

                            proftext = Entry(profile_frame,font=('arial',15),width=10,bg='lightblue')
                            proftext.insert(0, excel_activate['A' + str(each_cell)].value)
                            proftext.place(x=80,y=30)

                            type_label = Label(profile_frame,text="Type :",font=('arial',15),bg='lightblue')
                            type_var = StringVar()
                            type_list = ["BPI","BDO"," Metrobank","Gcash"]
                            type_combo = ttk.Combobox(profile_frame,textvariable=type_var,value=type_list,width=15,font=('arial'))

                            type_label.place(x=15,y=90)
                            type_combo.place(x=15,y=120)

                            balance = Label(profile_frame,text="Balance",bg='lightblue',font=('arial',20))
                            balance.place(x=15,y=170)

                            bal = Label(profile_frame,text=excel_activate['C' + str(each_cell)].value,bg='lightblue',font=('arial',25))
                            bal.place(x=125,y=165)

                            delbtn = Button(profile_frame,text="Delete",font=('arial',15),bg='lightblue',command=lambda:delete())
                            delbtn.place(x=15,y=250)

                            savebtn = Button(profile_frame,text="Save",font=('arial',15),bg='lightblue',command=lambda:save())
                            savebtn.place(x=150,y=250)

                            update_combobox_value()
                            profile.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(bank, profile))
                            profile.mainloop()
                                                            
                        else:
                            messagebox.showerror('Error','Wrong User ID')

                profilem_frame = Frame(profilem,width=300,height=200,borderwidth=20,relief='groove')
                profilem_frame.place(x=0,y=0)

                profilem_ent = Entry(profilem_frame,width=13,font=("arial",20))
                profilem_ent.insert(0, 'Enter UserID')
                profilem_ent.bind('<FocusIn>', on_entry_click1)
                profilem_ent.place(x=10,y=10)

                profilem_btn = Button(profilem_frame,text="Enter",font=('arial',20),command=lambda:profilem_function(),bg='lightblue')
                profilem_btn.place(x=10,y=50)

                profilem.mainloop()

            bank.mainloop()
        else:
            messagebox.showerror("Notification", "Wrong Entry!\nTry Again!")

def register():
    groot = Toplevel()
    groot.geometry("500x500")
    groot.title('REGISTER USER')
    root.withdraw()

    def go_back_to_main(root, groot):
        groot.destroy()
        root.deiconify()
        messagebox.showinfo("LogIN", "Welcome Back To LogIn")

    reg_frame = Frame(groot,width=500,height=500,borderwidth=28,relief='groove',bg='lightblue')
    reg_frame.place(x=0,y=0)

    register_t = Label(reg_frame,text='Register',font=('arial',25,'underline'),bg='lightblue')
    register_t.place(x=10,y=10)

    cut_label = Label(reg_frame,text='____________________________',font=('arial',20),bg='lightblue')
    cut_label.place(x=10,y=50)

    register_lbl = Label(reg_frame,text='Enter Your Desire User-ID',font=('arial',20),bg='lightblue')
    register_lbl.place(x=10,y=100)   

    userreg = Entry(reg_frame,width=20,font=('arial',20))
    userreg.place(x=10,y=160)

    pass_lbl = Label(reg_frame,text='Enter Your Desire Password',font=('arial',20),bg='lightblue')
    pass_lbl.place(x=10,y=210)   

    passreg = Entry(reg_frame,width=20,font=('arial',20),show="•")
    passreg.place(x=10,y=260)

    cut_label = Label(reg_frame,text='____________________________',font=('arial',20),bg='lightblue')
    cut_label.place(x=10,y=300)

    login_button = Button(reg_frame,text='REGISTER',font=('arial',20),bg='lightblue',width=25,borderwidth=5,command=lambda:register())
    login_button.place(x=10,y=350)

    def register():
        Found = False
        user = userreg.get()
        password = passreg.get()
        
        if user == "" and password == "":
            messagebox.showinfo("ERROR", "FILL ALL ENTRIES")
        else:
            for each_cell in range(2, excel_activate.max_row + 1):
                if user == excel_activate['A' + str(each_cell)].value:
                    Found = True
                    break
                else:
                    Found = False
            if Found:
                messagebox.showerror("ERROR", "Account Exist")
            else:
                lastrow = str(excel_activate.max_row + 1)
                excel_activate['A' + lastrow] = user
                excel_activate['B' + lastrow] = password
                excel_con.save('accounts.xlsx')
                messagebox.showinfo("SUCCESS", "Account Created")
                groot.destroy()
                root.deiconify()
    
    groot.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, groot))
    groot.mainloop()

login_frame = Frame(root,width=500,height=600,borderwidth=28,relief='groove',bg='lightblue')
login_frame.place(x=0,y=0)

login_label = Label(login_frame,text='Log-IN',font=('arial',30,'underline'),bg='lightblue')
login_label.place(x=10,y=30)

cut_label = Label(login_frame,text='____________________________',font=('arial',20),bg='lightblue')
cut_label.place(x=10,y=80)

user_label = Label(login_frame,text='User ID',font=('arial',30,'underline'),bg='lightblue')
user_label.place(x=10,y=150)

username = Entry(login_frame,width=18,font=('arial',30))
username.place(x=10,y=200)

pass_label = Label(login_frame,text='PassWord',font=('arial',30,'underline'),bg='lightblue')
pass_label.place(x=10,y=250)

password = Entry(login_frame,width=18,font=('arial',30),show="•")
password.place(x=10,y=300)

login_button = Button(login_frame,text='LOG IN',font=('arial',20),bg='lightblue',width=24,borderwidth=5,command=lambda:loginfunction())
login_button.place(x=10,y=370)

cut_label = Label(login_frame,text='____________________________',font=('arial',20),bg='lightblue')
cut_label.place(x=10,y=430)

register_label = Label(login_frame,text='First Time User?',font=('arial',10),bg='lightblue')
register_label.place(x=150,y=485)

reg_button = Button(login_frame,text='Register',font=('arial',10,'underline'),bg='lightblue',relief='sunken',command=lambda:register())
reg_button.place(x=255,y=480)

root.mainloop()